from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pyreadstat
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
import pandas as pd
from PySide6.QtWidgets import (
    QMessageBox,
)
from PySide6.QtCore import QThread, Signal


def _total(col: str, df) -> int:
    try:
        return len(df[col])
    except Exception:
        return 0


def _no_vacios(col: str, df) -> int:
    try:
        return df[col].count()
    except Exception:
        return 0


def _vacios(col: str, df) -> int:
    try:
        return df[col].isna().sum()
    except Exception:
        return 0

def _tipo_variable(col: str, df) -> str:
    try:
        s = df[col]

        if pd.api.types.is_integer_dtype(s):
            return "Entera"
        elif pd.api.types.is_float_dtype(s):
            return "Numérica"
        elif pd.api.types.is_string_dtype(s) or s.dtype == object:
            return "Texto"
        elif pd.api.types.is_datetime64_any_dtype(s):
            return "Fecha"
        elif pd.api.types.is_bool_dtype(s):
            return "Lógica"
        else:
            return "Otro"
    except Exception:
        return ""


def _build_rows_for_sav(path: str) -> list[dict]:
    """Lee un SAV y devuelve lista de dicts con los campos del diccionario."""
    df, meta = pyreadstat.read_sav(path)  # necesitamos df para min/max y errores
    col_labels = meta.column_names_to_labels or {}
    rows = []
    for i, col in enumerate(meta.column_names, start=1):
        tipo= _tipo_variable(col, df)
        etiqueta    = col_labels.get(col, "")

        total = _total(col, df)
        no_vacios = _no_vacios(col, df)
        vacios=_vacios(col, df)
        rows.append({
            "N":                          i,
            "TIPO DE VARIABLE":           tipo,
            "VARIABLE":           col,
            "PREGUNTA": etiqueta,
            "TOTAL":total,
            "NO VACIOS":no_vacios,
            "VACIOS":vacios,
        })
    return rows

# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_HEADER_COLS = [
    "Nº",
    "TIPO DE VARIABLE",
    "VARIABLE",
    "PREGUNTA",
    "TOTAL",
    "NO VACIOS",
    "VACIOS"
]

_COL_WIDTHS = [6, 28, 55, 55, 20, 11,38,8]

_FILL_HEADER = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_FILL_TITLE  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
_FILL_ALT    = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")

_FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_FONT_TITLE  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_FONT_DATA   = Font(name="Arial", size=9)

_THIN = Side(border_style="thin", color="B8CCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _write_sheet(ws, sav_name: str, rows: list[dict]) -> None:
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # Fila 1: título
    ws.merge_cells(f"A1:{get_column_letter(len(_HEADER_COLS))}1")
    cell = ws["A1"]
    cell.value = f"Revision de Valores — {sav_name.upper()}"
    cell.font = _FONT_TITLE
    cell.fill = _FILL_TITLE
    cell.alignment = _ALIGN_CENTER

    # Fila 2: encabezados
    for col_idx, (header, width) in enumerate(zip(_HEADER_COLS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filas de datos
    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 30
        #fill = _FILL_ALT if r_idx % 2 == 0 else None
        values = [
            row["N"],
            row["TIPO DE VARIABLE"],
            row["VARIABLE"],
            row["PREGUNTA"],
            row["TOTAL"],
            row["NO VACIOS"],
            row["VACIOS"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = _FONT_DATA
            cell.border = _BORDER
            cell.alignment = _ALIGN_CENTER if col_idx in (1, 5, 6) else _ALIGN_LEFT
            # if fill:
            #     cell.fill = fill
        max_lineas = 1
        anchos = {1: 5, 2: 20}
        for col_idx, val in enumerate(values, start=1):
            if val:
                texto = str(val)
                ancho = anchos.get(col_idx, 20)
                lineas = sum(
                    max(1, len(line) // ancho + 1)
                    for line in texto.split("\n")
                )
                max_lineas = max(max_lineas, lineas)
        ws.row_dimensions[r_idx].height = max_lineas * 15
    # Inmovilizar encabezados
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

def calcular_altura(ws, r_idx, ancho_col=30):
    max_lineas = 1
    for cell in ws[r_idx]:
        if cell.value:
            texto = str(cell.value)
            lineas = sum(
                max(1, len(line) // ancho_col + 1)
                for line in texto.split("\n")
            )
            max_lineas = max(max_lineas, lineas)
    ws.row_dimensions[r_idx].height = max_lineas * 15
def generar_revision_excel(
    carpeta_sav: str,
    ruta_salida: str,
    callback_progreso=None,
) -> str:
    """
    Lee todos los .sav de `carpeta_sav` y genera un .xlsx en `ruta_salida`.
    Devuelve la ruta final del archivo.
    callback_progreso(pct: int) se llama con 0‒100.
    """
    sav_files = sorted(Path(carpeta_sav).glob("*.sav")) + sorted(
        Path(carpeta_sav).glob("*.SAV")
    )
    sav_files = list(dict.fromkeys(sav_files))  # dedup

    if not sav_files:
        raise FileNotFoundError("No se encontraron archivos .sav en la carpeta.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    total = len(sav_files)
    for i, sav_path in enumerate(sav_files, start=1):
        if callback_progreso:
            callback_progreso(int((i - 1) / total * 90))

        rows = _build_rows_for_sav(str(sav_path))
        sheet_name = sav_path.stem[:31]  # Excel limita a 31 chars
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, sav_path.stem, rows)

    if callback_progreso:
        callback_progreso(95)

    out_path = Path(ruta_salida)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    if callback_progreso:
        callback_progreso(100)

    return str(out_path)







class _GenWorker(QThread):
    finished = Signal(str)
    error    = Signal(str)
    progress = Signal(int)

    def __init__(self, carpeta: str, salida: str):
        super().__init__()
        self.carpeta = carpeta
        self.salida  = salida

    def run(self):
        try:
            path = generar_revision_excel(
                self.carpeta,
                self.salida,
                callback_progreso=self.progress.emit,
            )
            self.finished.emit(path)
        except Exception as exc:
            self.error.emit(str(exc))


class Revisor:

    def __init__(self, parent=None):
        self._parent = parent
        self._worker: _GenWorker | None = None

    def ejecutar(self, carpeta: str, salida: str):
        self._worker = _GenWorker(carpeta, salida)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_finished(self, path: str):
        QMessageBox.information(
            self._parent, "¡Listo!",
            f" generado correctamente:\n\n{path}"
        )

    def _on_error(self, msg: str):
        QMessageBox.critical(self._parent, "Error", msg)