from __future__ import annotations

import re
from pathlib import Path
import string

import openpyxl
import pyreadstat
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
import pandas as pd
from PySide6.QtWidgets import (
    QMessageBox,
)
from PySide6.QtCore import QThread, Signal
import json

def comparar_campos_comunes(columnas: list[dict]) -> list[dict]:

    if len(columnas) < 2:
        raise ValueError("Se necesitan al menos 2 archivos para comparar")

    archivos = [entrada["ARCHIVO"] for entrada in columnas]
    NO_PRESENTE = "El campo no esta presente ."
    SIN_ETIQUETA = "Sin etiqueta."
    AUSENCIA = "__AUSENCIA__"  # marcador interno solo para "campo no existe"

    # campo -> {archivo: etiqueta}
    etiquetas_por_campo = {}
    for entrada in columnas:
        archivo = entrada["ARCHIVO"]
        for var in entrada["VARIABLES"]:
            campo = var["CAMPO"]
            etiquetas_por_campo.setdefault(campo, {})[archivo] = var["ETIQUETA"]

    def normalizar(v):
        """Trata None, '' o espacios en blanco como 'sin etiqueta'."""
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return SIN_ETIQUETA
        return v

    def clave_comparable(v):
        """
        Solo NO_PRESENTE (campo ausente del archivo) se excluye de la
        comparación. SIN_ETIQUETA SÍ es un valor comparable: un campo
        presente sin etiqueta es distinto de uno presente con etiqueta.
        """
        if v == NO_PRESENTE:
            return AUSENCIA
        return normalizar(v)

    resultado = []
    for campo, por_archivo in etiquetas_por_campo.items():
        # Debe aparecer en al menos 2 archivos para que tenga sentido comparar
        if len(por_archivo) < 2:
            continue

        # Valor crudo por archivo (para mostrar), marcando ausencia del campo
        etiquetas_crudas = {
            archivo: por_archivo.get(archivo, NO_PRESENTE) for archivo in archivos
        }

        claves = [clave_comparable(v) for v in etiquetas_crudas.values()]

        # Excluimos solo AUSENCIA (campo no presente en ese archivo);
        # SIN_ETIQUETA se queda como valor comparable
        claves_reales = {c for c in claves if c != AUSENCIA}

        # Si no hay ningún archivo donde el campo exista, o todos los
        # que existen coinciden (misma etiqueta, o todos sin etiqueta),
        # no hay diferencia que reportar
        if len(claves_reales) <= 1:
            continue

        fila = {"CAMPO": campo}
        for archivo in archivos:
            valor = etiquetas_crudas[archivo]
            if valor == NO_PRESENTE:
                fila[f"ETIQUETA {archivo}"] = NO_PRESENTE
            else:
                fila[f"ETIQUETA {archivo}"] = normalizar(valor)

        resultado.append(fila)

    return resultado
def comparar_valores_comunes(columnas: list[dict]) -> list[dict]:

    if len(columnas) < 2:
        raise ValueError("Se necesitan al menos 2 archivos para comparar")

    archivos = [entrada["ARCHIVO"] for entrada in columnas]
    NO_PRESENTE = "El campo no esta presente."
    SIN_ETIQUETAS = "Sin valores."
    AUSENCIA = "__AUSENCIA__"  # marcador interno solo para "campo no existe"

    # campo -> {archivo: etiquetas_de_valores}
    valores_por_campo = {}
    for entrada in columnas:
        archivo = entrada["ARCHIVO"]
        for var in entrada["VARIABLES"]:
            campo = var["CAMPO"]
            valores_por_campo.setdefault(campo, {})[archivo] = var["ETIQUETAS DE VALORES"]

    def normalizar(v):
        """Trata None y {} (o vacío) como 'sin etiquetas'."""
        if v is None or v == {} or v == "":
            return SIN_ETIQUETAS
        return v

    def clave_comparable(v):
        """
        Solo NO_PRESENTE (campo ausente del archivo) se excluye de la
        comparación. SIN_ETIQUETAS SÍ es un valor comparable: un campo
        presente sin etiquetas es distinto de uno presente con etiquetas.
        """
        if v == NO_PRESENTE:
            return AUSENCIA
        v_norm = normalizar(v)
        if v_norm == SIN_ETIQUETAS:
            return SIN_ETIQUETAS
        return json.dumps(v_norm, sort_keys=True, ensure_ascii=False)

    resultado = []
    for campo, por_archivo in valores_por_campo.items():
        if len(por_archivo) < 2:
            continue

        # Valor crudo por archivo (para mostrar), marcando ausencia de campo
        valores_crudos = {
            archivo: por_archivo.get(archivo, NO_PRESENTE) for archivo in archivos
        }

        claves = [clave_comparable(v) for v in valores_crudos.values()]

        # Excluimos solo AUSENCIA (campo no presente en ese archivo);
        # SIN_ETIQUETAS se queda como valor comparable
        claves_reales = {c for c in claves if c != AUSENCIA}

        # Si no hay ningún archivo donde el campo exista, o todos los
        # que existen coinciden (mismas etiquetas, o todos sin etiquetas),
        # no hay diferencia que reportar
        if len(claves_reales) <= 1:
            continue

        fila = {"CAMPO": campo}
        for archivo in archivos:
            valor = valores_crudos[archivo]
            if valor == NO_PRESENTE:
                fila[f"VALORES {archivo}"] = NO_PRESENTE
            elif normalizar(valor) == SIN_ETIQUETAS:
                fila[f"VALORES {archivo}"] = SIN_ETIQUETAS
            else:
                fila[f"VALORES {archivo}"] = valor

        resultado.append(fila)

    return resultado
def _fmt_valores(val_dict: dict) -> str:
    """Convierte {1.0: 'Sí', 2.0: 'No'} → '1. Sí\n2. No'"""
    if not val_dict:
        return ""
    partes = []
    for k, v in sorted(val_dict.items()):
        try:
            key_str = str(int(float(k)))
        except (ValueError, TypeError):
            key_str = str(k)
        partes.append(f"{key_str}. {v}")
    return "\n".join(partes)

def _build_rows_for_sav(path: Path) -> dict:
    path = Path(path)
    df, meta = pyreadstat.read_sav(path)

    col_labels = meta.column_names_to_labels or {}
    val_labels = meta.variable_value_labels or {}

    variables = []

    for col in meta.column_names:
        variables.append({
            "CAMPO": col,
            "ETIQUETA": col_labels.get(col, ""),
            "ETIQUETAS DE VALORES": _fmt_valores(val_labels.get(col, {})),
        })

    return {
        "ARCHIVO": path.name,
        "VARIABLES": variables
    }
_HEADER_COLS = [
    "CAMPO",
    "ETIQUETA A",
    "ETIQUETA B",
]

_COL_WIDTHS = [6, 28, 55, 55, 20, 11,38,8,20]

_FILL_HEADER = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_FILL_TITLE  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
_FILL_ALT    = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")

_FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_FONT_TITLE  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_FONT_DATA   = Font(name="Arial", size=9)

_THIN = Side(border_style="thin", color="B8CCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _write_sheet(ws, sav_name: str, rows: list[dict]) -> None:
    if not rows:
        return

    # Headers dinámicos a partir de las keys del primer row
    headers = list(rows[0].keys())
    col_widths = [30 if h == "Campo común" else 12 if h == "ETIQUETA_IGUAL" else 25 for h in headers]

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # # Fila 1: título
    # ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    # cell = ws["A1"]
    # cell.value = f"Revision de Valores — {sav_name.upper()}"
    # cell.font = _FONT_TITLE
    # cell.fill = _FILL_TITLE
    # cell.alignment = _ALIGN_CENTER

    # Fila 2: encabezados
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filas de datos
    for r_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 30
        values = [row[h] for h in headers]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = _FONT_DATA
            cell.border = _BORDER
            # Centra "Campo común" y "ETIQUETA_IGUAL", el resto alineado a la izquierda
            header = headers[col_idx - 1]
            cell.alignment = _ALIGN_CENTER if header in ("Campo común", "ETIQUETA_IGUAL") else _ALIGN_LEFT

        max_lineas = 1
        for col_idx, val in enumerate(values, start=1):
            if val:
                texto = str(val)
                ancho = col_widths[col_idx - 1]
                lineas = sum(
                    max(1, len(line) // ancho + 1)
                    for line in texto.split("\n")
                )
                max_lineas = max(max_lineas, lineas)
        ws.row_dimensions[r_idx].height = max_lineas * 15

    # Inmovilizar encabezados
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

def calcular_altura(ws, r_idx, ancho_col=30):
    max_lineas = 1
    for cell in ws[r_idx]:
        if cell.value:
            texto = str(cell.value)
            lineas = sum(
                max(1, len(line) // ancho_col + 1)
                for line in texto.split("\n")
            )
            max_lineas = max(max_lineas, lineas)
    ws.row_dimensions[r_idx].height = max_lineas * 15
def generar_revision_excel(
    carpeta_sav: str,
    ruta_salida: str,
    callback_progreso=None,
) -> str:
    """
    Lee todos los .sav de `carpeta_sav` y genera un .xlsx en `ruta_salida`.
    Devuelve la ruta final del archivo.
    callback_progreso(pct: int) se llama con 0‒100.
    """
    sav_files = sorted(Path(carpeta_sav).glob("*.sav")) + sorted(
        Path(carpeta_sav).glob("*.SAV")
    )
    sav_files = list(dict.fromkeys(sav_files))  # dedup

    if not sav_files:
        raise FileNotFoundError("No se encontraron archivos .sav en la carpeta.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto
    columnas = []
    total = len(sav_files)
    for i, sav_path in enumerate(sav_files, start=1):
        if callback_progreso:
            callback_progreso(int((i - 1) / total * 90))

        rows = _build_rows_for_sav(sav_path)
        columnas.append(rows)

    filas = comparar_campos_comunes(columnas)
    sheet_name = 'REVISION ETIQUETAS'
    ws = wb.create_sheet(title=sheet_name)
  
    _write_sheet(ws, sav_path.stem, filas)

    filas_valores = comparar_valores_comunes(columnas)
    sheet_name_valores = 'REVISION VALORES'
    ws2 = wb.create_sheet(title=sheet_name_valores)

    _write_sheet(ws2, sav_path.stem, filas_valores)

    if callback_progreso:
        callback_progreso(95)

    out_path = Path(ruta_salida)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    if callback_progreso:
        callback_progreso(100)

    return str(out_path)
class _GenWorker(QThread):
    finished = Signal(str)
    error    = Signal(str)
    progress = Signal(int)

    def __init__(self, carpeta: str, salida: str):
        super().__init__()
        self.carpeta = carpeta
        self.salida  = salida

    def run(self):
        try:
            path = generar_revision_excel(
                self.carpeta,
                self.salida,
                callback_progreso=self.progress.emit,
            )
            self.finished.emit(path)
        except Exception as exc:
            self.error.emit(str(exc))


class RevisorEtiquetas:

    def __init__(self, parent=None):
        self._parent = parent
        self._worker: _GenWorker | None = None

    def ejecutar(self, carpeta: str, salida: str):
        self._worker = _GenWorker(carpeta, salida)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_finished(self, path: str):
        QMessageBox.information(
            self._parent, "¡Listo!",
            f" generado correctamente:\n\n{path}"
        )

    def _on_error(self, msg: str):
        QMessageBox.critical(self._parent, "Error", msg)