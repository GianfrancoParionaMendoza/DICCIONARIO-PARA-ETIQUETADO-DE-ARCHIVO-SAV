from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pyreadstat
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import (
    QMessageBox,
)
from PySide6.QtCore import QThread, Signal


def _eval_longitud(
    col: str,
    df,
    tipo: str,
    longitud_declarada: str,
    valores_dict: dict | None = None,
) -> tuple[str, str]:
    try:
        serie = df[col].dropna()
        if serie.empty and not valores_dict:
            return longitud_declarada, ""

        if tipo == "A":
            long_real = int(serie.astype(str).str.len().max()) if not serie.empty else 0

        elif tipo == "N":
            def num_len(v):
                try:
                    if float(v) == int(float(v)):
                        return len(str(int(float(v))))
                    return len(str(v))
                except (TypeError, ValueError):
                    return len(str(v))

            largos = list(serie.apply(num_len)) if not serie.empty else []

            # Las claves de las etiquetas (ej. 0, 99) pueden no estar presentes
            # en la data cargada, pero sí forman parte del rango declarado
            # (ver _fmt_rango_variacion). La longitud debe contemplarlas también.
            if valores_dict:
                largos.extend(num_len(k) for k in valores_dict.keys())

            long_real = max(largos) if largos else 0

        else:
            # AN (fecha/datetime): devolver la declarada sin evaluar
            return longitud_declarada, ""

        # Advertencia si la real supera la declarada
        adv = ""
        try:
            if long_real > int(longitud_declarada):
                adv = f"longitud real ({long_real}) > declarada ({longitud_declarada})"
        except (ValueError, TypeError):
            pass

        return str(long_real), adv  # ← siempre la real

    except Exception:
        return longitud_declarada, ""
def _parse_fmt(fmt: str) -> tuple[str, str, str]:
    """Devuelve (TIPO, LONGITUD, DECIMALES) desde formato SPSS."""
    if not fmt:
        return "", "", ""
    m = re.match(r"[Aa](\d+)", fmt)
    if m:
        return "A", m.group(1), ""
    m = re.match(r"[Ff](\d+)\.(\d+)", fmt)
    if m:
        return "N", m.group(1), m.group(2)
    m = re.match(r"[Nn](\d+)", fmt)
    if m:
        return "N", m.group(1), "0"
    m = re.match(r"E(\d+)\.(\d+)", fmt, re.IGNORECASE)
    if m:
        return "N", m.group(1), m.group(2)
    m = re.match(r"(?:DOLLAR|COMMA|DOT|PCT)(\d+)\.?(\d+)?", fmt, re.IGNORECASE)
    if m:
        return "N", m.group(1) or "8", m.group(2) or "0"
    m = re.match(r"(?:DATETIME|DTIME)(\d*)", fmt, re.IGNORECASE)
    if m:
        return "AN", m.group(1) or "20", ""
    m = re.match(r"(?:ADATE|EDATE|JDATE|SDATE|QYR|MOYR|WKYR|DATE|TIME)(\d*)", fmt, re.IGNORECASE)
    if m:
        return "AN", m.group(1) or "8", ""
    return "N", "8", "0"


def _fmt_valores(val_dict: dict) -> str:
    """Convierte {1.0: 'Sí', 2.0: 'No'} → '1. Sí\n2. No'"""
    if not val_dict:
        return ""
    partes = []
    for k, v in sorted(val_dict.items()):
        try:
            key_str = str(int(float(k)))
        except (ValueError, TypeError):
            key_str = str(k)
        partes.append(f"{key_str}. {v}")
    return "\n".join(partes)

def _fmt_rango_variacion(col: str, df, val_labels: dict) -> tuple[str, str]:
    valores_dict = val_labels.get(col, {})
    tiene_etiquetas = bool(valores_dict)

    try:
        serie = df[col].dropna()
    except Exception:
        return "", ""

    if serie.empty:
        return "", ""

    if not tiene_etiquetas:
        try:
            min_val = int(serie.min()) if serie.min() == int(serie.min()) else serie.min()
            max_val = int(serie.max()) if serie.max() == int(serie.max()) else serie.max()
            rango = f"[{min_val} : {max_val}]"
        except (TypeError, ValueError):
            rango = ""
        return rango, ""

    # Claves etiquetadas
    claves_set = set()
    for k in valores_dict.keys():
        try:
            claves_set.add(float(k))
        except (TypeError, ValueError):
            pass

    # Valores reales presentes en la data
    valores_reales = set()
    for v in serie.unique():
        try:
            valores_reales.add(float(v))
        except (TypeError, ValueError):
            pass

    # El rango contempla AMBOS: etiquetas + valores reales (ej. un 0 sin etiqueta)
    todos = claves_set | valores_reales
    try:
        min_v = min(todos)
        max_v = max(todos)
        min_v = int(min_v) if min_v == int(min_v) else min_v
        max_v = int(max_v) if max_v == int(max_v) else max_v
        rango = f"[{min_v} : {max_v}]"
    except Exception:
        rango = ""

    # Valores que por convención del sistema representan "nulo / no aplica"
    # y por eso NUNCA requieren tener una etiqueta propia en el diccionario.
    _VALORES_NULOS_CONVENCION = {0.0}

    # Errores: valores reales sin etiqueta, EXCLUYENDO los nulos por convención
    fuera = []
    for v in serie.unique():
        try:
            v_float = float(v)
            if v_float not in claves_set and v_float not in _VALORES_NULOS_CONVENCION:
                fuera.append(int(v) if v_float == int(v_float) else v_float)
        except (TypeError, ValueError):
            if v not in claves_set:
                fuera.append(v)

    fuera.sort(key=lambda x: str(x))
    error = f"{fuera}" if fuera else ""

    return rango, error

def _build_rows_for_sav(path: str) -> list[dict]:
    """Lee un SAV y devuelve lista de dicts con los campos del diccionario."""
    df, meta = pyreadstat.read_sav(path)  # necesitamos df para min/max y errores

    col_labels = meta.column_names_to_labels or {}
    val_labels = meta.variable_value_labels or {}
    formats    = meta.original_variable_types or {}

    rows = []
    for i, col in enumerate(meta.column_names, start=1):
        etiqueta    = col_labels.get(col, "")
        valores_dict = val_labels.get(col, {})
        valores_str = _fmt_valores(valores_dict)
        fmt         = formats.get(col, "")
        tipo, longitud, _ = _parse_fmt(fmt)
        longitudeva, adv_longitud = _eval_longitud(col, df, tipo, longitud)
        rango, error = _fmt_rango_variacion(col, df, val_labels)
        # Combinar advertencias en ERROR
        errores = []
        if adv_longitud:
            errores.append(adv_longitud)
        if error:
            errores.append(error)
        error_final = "  |  ".join(errores)

        rows.append({
            "N":                          i,
            "NOMBRE DEL CAMPO":           col,
            "DESCRIPCIÓN DE LAS VARIABLES": etiqueta,
            "ETIQUETAS DE VALORES":       valores_str,
            "TIPO DE CARÁCTER":           tipo,
            "LONGITUD":                   longitudeva,
            "RANGO DE VARIACION":         rango,
            "ERROR":                      error_final,
        })
    return rows

# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_HEADER_COLS = [
    "Nº",
    "NOMBRE DEL CAMPO",
    "DESCRIPCIÓN DE LAS VARIABLES",
    "ETIQUETAS DE VALORES",
    "TIPO DE CARÁCTER",
    "LONGITUD",
    "RANGO DE VARIACION",
    "ERROR",
]

_COL_WIDTHS = [6, 28, 55, 55, 16, 10,10,10]

_FILL_HEADER = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
_FILL_TITLE  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
_FILL_ALT    = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")

_FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_FONT_TITLE  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_FONT_DATA   = Font(name="Arial", size=9)

_THIN = Side(border_style="thin", color="B8CCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _write_sheet(ws, sav_name: str, rows: list[dict]) -> None:
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # Fila 1: título
    ws.merge_cells(f"A1:{get_column_letter(len(_HEADER_COLS))}1")
    cell = ws["A1"]
    cell.value = f"DICCIONARIO DE DATOS — {sav_name.upper()}"
    cell.font = _FONT_TITLE
    cell.fill = _FILL_TITLE
    cell.alignment = _ALIGN_CENTER

    # Fila 2: encabezados
    for col_idx, (header, width) in enumerate(zip(_HEADER_COLS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filas de datos
    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 30
        fill = _FILL_ALT if r_idx % 2 == 0 else None
        values = [
            row["N"],
            row["NOMBRE DEL CAMPO"],
            row["DESCRIPCIÓN DE LAS VARIABLES"],
            row["ETIQUETAS DE VALORES"],
            row["TIPO DE CARÁCTER"],
            row["LONGITUD"],
            row["RANGO DE VARIACION"],
            row["ERROR"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font = _FONT_DATA
            cell.border = _BORDER
            cell.alignment = _ALIGN_CENTER if col_idx in (1, 5, 6) else _ALIGN_LEFT
            if fill:
                cell.fill = fill

    # Inmovilizar encabezados
    ws.freeze_panes = "A3"


def generar_diccionario_excel(
    carpeta_sav: str,
    ruta_salida: str,
    callback_progreso=None,
) -> str:
    """
    Lee todos los .sav de `carpeta_sav` y genera un .xlsx en `ruta_salida`.
    Devuelve la ruta final del archivo.
    callback_progreso(pct: int) se llama con 0‒100.
    """
    sav_files = sorted(Path(carpeta_sav).glob("*.sav")) + sorted(
        Path(carpeta_sav).glob("*.SAV")
    )
    sav_files = list(dict.fromkeys(sav_files))  # dedup

    if not sav_files:
        raise FileNotFoundError("No se encontraron archivos .sav en la carpeta.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    total = len(sav_files)
    for i, sav_path in enumerate(sav_files, start=1):
        if callback_progreso:
            callback_progreso(int((i - 1) / total * 90))

        rows = _build_rows_for_sav(str(sav_path))
        sheet_name = sav_path.stem[:31]  # Excel limita a 31 chars
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, sav_path.stem, rows)

    if callback_progreso:
        callback_progreso(95)

    out_path = Path(ruta_salida)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    if callback_progreso:
        callback_progreso(100)

    return str(out_path)







class _GenWorker(QThread):
    finished = Signal(str)
    error    = Signal(str)
    progress = Signal(int)

    def __init__(self, carpeta: str, salida: str):
        super().__init__()
        self.carpeta = carpeta
        self.salida  = salida

    def run(self):
        try:
            path = generar_diccionario_excel(
                self.carpeta,
                self.salida,
                callback_progreso=self.progress.emit,
            )
            self.finished.emit(path)
        except Exception as exc:
            self.error.emit(str(exc))


class GeneradorDiccionario:

    def __init__(self, parent=None):
        self._parent = parent
        self._worker: _GenWorker | None = None

    def ejecutar(self, carpeta: str, salida: str):
        self._worker = _GenWorker(carpeta, salida)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_finished(self, path: str):
        QMessageBox.information(
            self._parent, "¡Listo!",
            f"Diccionario generado correctamente:\n\n{path}"
        )

    def _on_error(self, msg: str):
        QMessageBox.critical(self._parent, "Error", msg)