from __future__ import annotations

import re
import openpyxl
from pathlib import Path

import pyreadstat
from PySide6.QtCore import QThread, Signal
from PySide6.QtWidgets import QMessageBox

from .loader import load_sav
from datetime import datetime
import pandas as pd
import shutil
def _adaptar_keys(val_dict: dict, col_dtype: str) -> dict:
    """Retorna SIEMPRE un dict nuevo, nunca modifica el original."""
    col_dtype = col_dtype.lower()
    if col_dtype in ("object", "string", "str") or "str" in col_dtype:
        return {
            str(int(k)) if isinstance(k, float) and k == int(k) else str(k): v
            for k, v in val_dict.items()
        }
    elif "float" in col_dtype:
        return {float(k): v for k, v in val_dict.items()}
    elif "int" in col_dtype:
        return {int(k): v for k, v in val_dict.items()}
    # fallback: copia limpia sin mutar el original
    return dict(val_dict)

class _ProcesarWorker(QThread):
    """Procesa todos los SAV de una carpeta en segundo plano."""
    progress      = Signal(int, str)
    archivo_listo = Signal(str, str)
    finished      = Signal(int, int)

    def __init__(self, insumo_path: Path, sav_paths: list[str], out_dir: Path,
                 excel_labels: dict, excel_values: dict, excel_fmt: dict):
        super().__init__()
        self.insumo_path  = insumo_path
        self.sav_paths    = sav_paths
        self.out_dir      = out_dir
        self.excel_labels = excel_labels
        self.excel_values = excel_values
        self.excel_fmt    = excel_fmt

    def run(self):
        total     = len(self.sav_paths)
        ok_count  = 0
        err_count = 0
        timestamp = datetime.now().strftime("%y%m%d_%H%M")

        lote_dir = self.out_dir / timestamp
        lote_dir.mkdir(parents=True, exist_ok=True)
        log = lote_dir / "insumo"
        shutil.copytree(str(self.insumo_path), str(log))

        for i, path_str in enumerate(self.sav_paths, start=1):
            nombre = Path(path_str).stem
            self.progress.emit(int((i - 1) / total * 100), nombre)
            try:
                self._procesar_uno(path_str, lote_dir)
                self.archivo_listo.emit(nombre, "ok")
                ok_count += 1
            except Exception as exc:
                self.archivo_listo.emit(nombre, str(exc))
                err_count += 1

        self.progress.emit(100, "")
        self.finished.emit(ok_count, err_count)

    def _procesar_uno(self, path_str: str, lote_dir: Path):
        sav_b     = load_sav(path_str)
        out_path  = lote_dir / f"{Path(path_str).stem}.sav"
        print(f"[DEBUG] entrada : {path_str}")
        print(f"[DEBUG] salida  : {out_path}")
        print(f"[DEBUG] existe  : {out_path.exists()}")  
       
        # ── Convertir a pandas UNA sola vez aquí ─────────────────────────────
        df_pandas = sav_b.df.to_pandas()
        

        col_labels: dict[str, str]  = {}
        val_labels: dict[str, dict] = {}
        var_fmt:    dict[str, str]  = {}
        for col, fmt in self.excel_fmt.items():
            if col not in df_pandas.columns:
                continue

            if fmt.startswith("F"):
                df_pandas[col] = pd.to_numeric(df_pandas[col], errors="coerce")

            elif fmt.startswith("A"):
                df_pandas[col] = df_pandas[col].astype("string")
                
        for col in sav_b.columns:
           
            if col in self.excel_labels:
                col_labels[col] = self.excel_labels[col]

            if col in self.excel_values:
                val_dict = self.excel_values.get(col, {})
                val_labels[col] = val_dict

            if col in self.excel_fmt:
                var_fmt[col] = self.excel_fmt[col]
            else:
                var_fmt[col] = _inferir_fmt_desde_sav(sav_b.df[col])
    
        pyreadstat.write_sav(
            df_pandas,
            str(out_path),
            variable_format=var_fmt,
            column_labels=col_labels,
            variable_value_labels=val_labels,
        )


# ─── Inferencia de formato desde los datos del SAV ───────────────────────────

def _inferir_fmt_desde_sav(serie) -> str:
    """
    Analiza los valores de una columna Polars (o pandas) y decide el formato SPSS:
      - Solo numéricos (o vacíos)  → F8.0
      - Alfanumérico / texto       → A200
      - Sin datos                  → A200
    """
    # Convertir a lista descartando nulos / NaN
    try:
        # Polars
        valores = [v for v in serie.to_list() if v is not None]
    except AttributeError:
        # pandas fallback
        import pandas as pd
        valores = [v for v in serie.dropna().tolist()]

    if not valores:
        return "A200"

    for v in valores:
        # Si algún valor tiene al menos una letra → cadena
        if re.search(r"[A-Za-záéíóúÁÉÍÓÚñÑ]", str(v)):
            return "A200"

    # Todos los valores son convertibles a número
    return "F8.0"


class Etiquetador:

    def __init__(self, parent=None):
        self._parent        = parent
        self._worker:       _ProcesarWorker | None = None
        self._excel_labels: dict[str, str]  = {}
        self._excel_values: dict[str, dict] = {}
        self._excel_fmt:    dict[str, str]  = {}

    def ejecutar(self, excel_path: str, carpeta_sav: str, carpeta_salida: str):
        try:
            self._cargar_excel(excel_path)
        except Exception as exc:
            QMessageBox.critical(self._parent, "Error al cargar Excel", str(exc))
            return

        if not self._excel_labels:
            QMessageBox.warning(
                self._parent, "Sin datos",
                "No se encontraron campos en el Excel. Verifica los nombres de columnas."
            )
            return

        sav_dir   = Path(carpeta_sav)
        visto:     set[str]  = set()
        sav_files: list[str] = []
        for p in sorted(sav_dir.iterdir()):
            if p.suffix.lower() == ".sav":
                key = p.name.lower()
                if key not in visto:
                    visto.add(key)
                    sav_files.append(str(p))
        if not sav_files:
            QMessageBox.warning(
                self._parent, "Sin archivos",
                f"No hay archivos .sav en:\n{carpeta_sav}"
            )
            return

        out_dir = Path(carpeta_salida)
        out_dir.mkdir(parents=True, exist_ok=True)

        self._worker = _ProcesarWorker(
            carpeta_sav,
            sav_files, out_dir,
            self._excel_labels, self._excel_values, self._excel_fmt,
        )
        self._worker.finished.connect(
            lambda ok, err: self._on_finalizado(ok, err, out_dir)
        )
        self._worker.start()

    def _cargar_excel(self, path_a: str):
        wb = openpyxl.load_workbook(path_a, data_only=True)

        self._excel_labels = {}
        self._excel_values = {}
        self._excel_fmt    = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            headers = {}
            for cell in ws[1]:
                if cell.value:
                    headers[str(cell.value).strip()] = cell.column - 1

            col_campo  = headers.get("NOMBRE DEL CAMPO")
            col_desc   = headers.get("DESCRIPCIÓN DE LAS VARIABLES")
            col_etiq   = headers.get("ETIQUETAS DE VALORES")
            col_longit = headers.get("LONGITUD")
            col_tipo   = headers.get("TIPO DE CARÁCTER")
            
            if col_campo is None or col_desc is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[col_campo] is None:
                    continue
                campo = str(row[col_campo]).strip()
                if not campo:
                    continue

                # Etiqueta de variable
                etiq = str(row[col_desc]).strip() if row[col_desc] is not None else ""
                self._excel_labels[campo] = etiq

                # Etiquetas de valores
                if col_etiq is not None and row[col_etiq] is not None:
                    self._excel_values[campo] = self._parse_excel_valores(str(row[col_etiq]))

                # Formato desde Excel (longitud + tipo)
                col_tipo_str = str(row[col_tipo]).strip().upper() if col_tipo is not None and row[col_tipo] is not None else ""
                longitud     = int(row[col_longit]) if col_longit is not None and row[col_longit] is not None else 0

                if col_tipo_str == "AN":
                    self._excel_fmt[campo] = f"A{longitud}"
                elif col_tipo_str == "N":
                    self._excel_fmt[campo] = f"F{longitud}.0" 
                # Si col_tipo_str está vacío o es otro valor, NO se registra en _excel_fmt
                # → el worker llamará a _inferir_fmt_desde_sav para ese campo

    def _on_finalizado(self, ok_count: int, err_count: int, out_dir: Path):
        QMessageBox.information(
            self._parent, "Exportación completa",
            f"Procesados: {ok_count + err_count}\n"
            f"Correctos:  {ok_count}\n"
            f"Errores:    {err_count}\n\n"
            f"Carpeta:\n{out_dir}"
        )

    @staticmethod
    def _parse_excel_valores(raw: str) -> dict:
        """
        Cubre todos los formatos del Excel real:
        '1. Urbana'         → número + punto + espacio (estándar)
        '1.Urbana'          → número + punto + sin espacio
        '0 Establecimiento' → número + solo espacio (sin punto)
        Mezclas de los anteriores en el mismo campo
        """
        result: dict = {}
        if not raw:
            return result

        patron = re.compile(
            r'^\s*'       # espacios iniciales opcionales
            r'(\d+)'      # ① clave numérica
            r'[.\s]'      # punto O espacio como separador
            r'\s*'        # espacios extra opcionales
            r'(.+)',      # ② etiqueta (resto de la línea)
            re.MULTILINE
        )

        for m in patron.finditer(raw):
            # ── Key: int cuando es entero exacto, float si tiene decimales ──
            raw_key = float(m.group(1))
            key = int(raw_key) if raw_key == int(raw_key) else raw_key

            lbl = m.group(2).strip()
            if lbl:
                result[key] = lbl
        
        return result
